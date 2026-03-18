from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File, Query
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from openpyxl import load_workbook

from app.db.database import get_db
from app.dependencies.auth import get_current_user
from app.models.lead import Lead
from app.models.destination import Destination
from app.models.pipeline import Pipeline, PipelineStage

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

LEAD_STATUS_OPTIONS = [
    ("novo", "Novo"),
    ("em_contato", "Em contato"),
    ("negociacao", "Negociação"),
    ("fechado", "Fechado"),
    ("perdido", "Perdido"),
]


def is_valid_status(status: str) -> bool:
    return status in [value for value, _ in LEAD_STATUS_OPTIONS]


def get_status_label(status: str) -> str:
    mapping = {value: label for value, label in LEAD_STATUS_OPTIONS}
    return mapping.get(status, status)


@router.get("/leads", response_class=HTMLResponse)
def list_leads(
    request: Request,
    status: str | None = Query(default=None),
    destination_id: int | None = Query(default=None),
    search: str | None = Query(default=None),
    travel_start: str | None = Query(default=None),
    travel_end: str | None = Query(default=None),
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    destinations = db.query(Destination).order_by(Destination.name.asc()).all()

    query = (
        db.query(Lead)
        .options(joinedload(Lead.destinations))
        .order_by(Lead.id.desc())
    )

    if status and is_valid_status(status):
        query = query.filter(Lead.status == status)

    if destination_id:
        query = query.join(Lead.destinations).filter(Destination.id == destination_id)

    if search and search.strip():
        search_value = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Lead.name.ilike(search_value),
                Lead.email.ilike(search_value),
                Lead.whatsapp.ilike(search_value)
            )
        )
        
    if travel_start:
        query = query.filter(Lead.travel_start_date >= travel_start)
    
    if travel_end:
        query = query.filter(Lead.travel_end_date <= travel_end)

    leads = query.all()

    return templates.TemplateResponse(
        "leads/list.html",
        {
            "request": request,
            "leads": leads,
            "current_user": current_user,
            "get_status_label": get_status_label,
            "status_options": LEAD_STATUS_OPTIONS,
            "destinations": destinations,
            "selected_status": status or "",
            "selected_destination_id": destination_id,
            "search_value": search or "",
            "travel_start": travel_start or "",
            "travel_end": travel_end or ""
        }
    )


@router.get("/leads/create", response_class=HTMLResponse)
def create_lead_page(
    request: Request,
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    destinations = db.query(Destination).order_by(Destination.name.asc()).all()
    pipelines = db.query(Pipeline).options(joinedload(Pipeline.stages)).all()

    return templates.TemplateResponse(
        "leads/create.html",
        {
            "request": request,
            "current_user": current_user,
            "destinations": destinations,
            "pipelines": pipelines,
            "status_options": LEAD_STATUS_OPTIONS,
            "error": None
        }
    )


@router.post("/leads/create", response_class=HTMLResponse)
def create_lead(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    whatsapp: str = Form(...),
    status: str = Form(...),
    destination_ids: list[int] = Form(...),
    travel_start_date: str = Form(...),
    travel_end_date: str = Form(...),
    pipeline_stage_id: int | None = Form(None),
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    destinations = db.query(Destination).order_by(Destination.name.asc()).all()

    name = name.strip()
    email = email.strip().lower()
    whatsapp = whatsapp.strip()
    status = status.strip()

    if not name or not email or not whatsapp:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Preencha todos os campos obrigatórios."
            }
        )
        
    existing_lead = db.query(Lead).filter(Lead.email == email).first()
    if existing_lead:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Já existe um lead com esse E-mail cadastrado."
            }
        )

    if not is_valid_status(status):
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione um status válido."
            }
        )

    if not destination_ids or len(destination_ids) == 0:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione pelo menos um destino."
            }
        )

    try:
        start_date = datetime.strptime(travel_start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(travel_end_date, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Datas inválidas."
            }
        )

    if end_date < start_date:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "A data final não pode ser menor que a data inicial."
            }
        )

    selected_destinations = (
        db.query(Destination)
        .filter(Destination.id.in_(destination_ids))
        .all()
    )

    if len(selected_destinations) == 0:
        return templates.TemplateResponse(
            "leads/create.html",
            {
                "request": request,
                "current_user": current_user,
                "destinations": destinations,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione pelo menos um destino válido."
            }
        )

    new_lead = Lead(
        name=name,
        email=email,
        whatsapp=whatsapp,
        status=status,
        travel_start_date=start_date,
        travel_end_date=end_date,
        pipeline_stage_id=pipeline_stage_id
    )

    new_lead.destinations = selected_destinations

    db.add(new_lead)
    db.commit()
    db.refresh(new_lead)

    return RedirectResponse(url="/leads", status_code=303)


@router.get("/leads/{lead_id}/edit", response_class=HTMLResponse)
def edit_lead_page(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    lead = (
        db.query(Lead)
        .options(joinedload(Lead.destinations))
        .filter(Lead.id == lead_id)
        .first()
    )

    if not lead:
        return RedirectResponse(url="/leads", status_code=303)

    destinations = db.query(Destination).order_by(Destination.name.asc()).all()
    pipelines = db.query(Pipeline).options(joinedload(Pipeline.stages)).all()
    selected_destination_ids = [destination.id for destination in lead.destinations]

    return templates.TemplateResponse(
        "leads/edit.html",
        {
            "request": request,
            "current_user": current_user,
            "lead": lead,
            "destinations": destinations,
            "pipelines": pipelines,
            "selected_destination_ids": selected_destination_ids,
            "status_options": LEAD_STATUS_OPTIONS,
            "error": None
        }
    )


@router.post("/leads/{lead_id}/edit", response_class=HTMLResponse)
def update_lead(
    lead_id: int,
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    whatsapp: str = Form(...),
    status: str = Form(...),
    destination_ids: list[int] = Form(...),
    travel_start_date: str = Form(...),
    travel_end_date: str = Form(...),
    pipeline_stage_id: int | None = Form(None),
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    destinations = db.query(Destination).order_by(Destination.name.asc()).all()

    if not lead:
        return RedirectResponse(url="/leads", status_code=303)

    name = name.strip()
    email = email.strip().lower()
    whatsapp = whatsapp.strip()
    status = status.strip()

    if not name or not email or not whatsapp:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Preencha todos os campos obrigatórios."
            }
        )
        
    existing_lead = db.query(Lead).filter(Lead.email == email, Lead.id != lead_id).first()
    if existing_lead:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Já existe outro lead com esse E-mail cadastrado."
            }
        )

    if not is_valid_status(status):
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione um status válido."
            }
        )

    if not destination_ids or len(destination_ids) == 0:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione pelo menos um destino."
            }
        )

    try:
        start_date = datetime.strptime(travel_start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(travel_end_date, "%Y-%m-%d").date()
    except ValueError:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Datas inválidas."
            }
        )

    if end_date < start_date:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "A data final não pode ser menor que a data inicial."
            }
        )

    selected_destinations = (
        db.query(Destination)
        .filter(Destination.id.in_(destination_ids))
        .all()
    )

    if len(selected_destinations) == 0:
        return templates.TemplateResponse(
            "leads/edit.html",
            {
                "request": request,
                "current_user": current_user,
                "lead": lead,
                "destinations": destinations,
                "selected_destination_ids": destination_ids,
                "status_options": LEAD_STATUS_OPTIONS,
                "error": "Selecione pelo menos um destino válido."
            }
        )

    lead.name = name
    lead.email = email
    lead.whatsapp = whatsapp
    lead.status = status
    lead.travel_start_date = start_date
    lead.travel_end_date = end_date
    lead.pipeline_stage_id = pipeline_stage_id
    lead.destinations = selected_destinations

    db.commit()

    return RedirectResponse(url="/leads", status_code=303)


@router.post("/leads/{lead_id}/delete")
def delete_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    try:
        get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    lead = db.query(Lead).filter(Lead.id == lead_id).first()

    if lead:
        db.delete(lead)
        db.commit()

    return RedirectResponse(url="/leads", status_code=303)


@router.get("/leads/import", response_class=HTMLResponse)
def import_leads_page(
    request: Request,
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    return templates.TemplateResponse(
        "leads/import.html",
        {
            "request": request,
            "current_user": current_user,
            "error": None,
            "success": None
        }
    )


@router.post("/leads/import", response_class=HTMLResponse)
async def import_leads_submit(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        current_user = get_current_user(request)
    except:
        return RedirectResponse(url="/login", status_code=303)

    if not file.filename or not file.filename.endswith(".xlsx"):
        return templates.TemplateResponse(
            "leads/import.html",
            {
                "request": request,
                "current_user": current_user,
                "error": "Envie um arquivo .xlsx",
                "success": None
            }
        )

    content = await file.read()
    workbook = load_workbook(filename=BytesIO(content))
    sheet = workbook.active

    imported_count = 0
    errors = []

    destination_map = {
        destination.name.strip().upper(): destination
        for destination in db.query(Destination).all()
    }

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            errors.append(f"Linha {row_index}: linha vazia.")
            continue

        if len(row) < 7:
            errors.append(f"Linha {row_index}: colunas insuficientes.")
            continue

        nome, email, whatsapp, status, destinos, data_ida, data_volta = row[:7]

        if not nome or not email or not whatsapp or not status or not destinos or not data_ida or not data_volta:
            errors.append(f"Linha {row_index}: campos obrigatórios ausentes.")
            continue

        nome = str(nome).strip()
        email = str(email).strip().lower()
        whatsapp = str(whatsapp).strip()
        status = str(status).strip()

        if not is_valid_status(status):
            errors.append(f"Linha {row_index}: status inválido.")
            continue

        raw_destinations = [item.strip().upper() for item in str(destinos).split(",") if item.strip()]
        matched_destinations = [destination_map[item] for item in raw_destinations if item in destination_map]

        if len(matched_destinations) == 0:
            errors.append(f"Linha {row_index}: nenhum destino válido encontrado.")
            continue

        try:
            if hasattr(data_ida, "date"):
                start_date = data_ida.date()
            else:
                start_date = datetime.strptime(str(data_ida), "%Y-%m-%d").date()

            if hasattr(data_volta, "date"):
                end_date = data_volta.date()
            else:
                end_date = datetime.strptime(str(data_volta), "%Y-%m-%d").date()
        except ValueError:
            errors.append(f"Linha {row_index}: datas inválidas.")
            continue

        if end_date < start_date:
            errors.append(f"Linha {row_index}: data final menor que data inicial.")
            continue

        new_lead = Lead(
            name=nome,
            email=email,
            whatsapp=whatsapp,
            status=status,
            travel_start_date=start_date,
            travel_end_date=end_date
        )
        new_lead.destinations = matched_destinations

        db.add(new_lead)
        imported_count += 1

    db.commit()

    success_message = f"{imported_count} lead(s) importado(s) com sucesso."

    if errors:
        success_message += f" {len(errors)} linha(s) com erro."

    return templates.TemplateResponse(
        "leads/import.html",
        {
            "request": request,
            "current_user": current_user,
            "error": errors if errors else None,
            "success": success_message
        }
    )